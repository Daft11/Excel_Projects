
import xlrd
import openpyxl
from openpyxl.workbook import Workbook

article_list = []
amount = []

# Забираем имя клиента из значений в таблице для имени файла


def file_name(file):
    wb_old = xlrd.open_workbook(file)
    sheet_active = wb_old.sheet_by_index(0)
    old_name = str(sheet_active.cell_value(8, 1))
    if ':' in old_name:
        old_name = old_name.split(':')
        old_name = old_name[0]
    # Проверяем имя на лишние символы
    #i = 0
    # while i != len(old_name):
    #    # Когда наткнется на ':' остановится
    #    if old_name[i] == ':':
    #        break
    #    else:
    #        i += 1
    # Забираем все символы до нежелательных двоеточий
    #name = old_name[:i]
    print(str(old_name))
    file_name_func = old_name.replace(
        ' ', '_')+'_№'+str(int(sheet_active.cell_value(7, 7)))+'.xlsx'
    return file_name_func

# Конвертация файла из xls в xlsx для дальнейшей работы с данными


def cvt_xls_to_xlsx(src_file_path, dst_file_path):
    book_xls = xlrd.open_workbook(src_file_path)
    book_xlsx = Workbook()

    sheet_names = book_xls.sheet_names()
    for sheet_index, sheet_name in enumerate(sheet_names):
        sheet_xls = book_xls.sheet_by_name(sheet_name)
        if sheet_index == 0:
            sheet_xlsx = book_xlsx.active
            sheet_xlsx.title = sheet_name
        else:
            sheet_xlsx = book_xlsx.create_sheet(title=sheet_name)

        for row in range(15, sheet_xls.nrows):
            for col in range(0, sheet_xls.ncols):
                sheet_xlsx.cell(row=row+1, column=col +
                                1).value = sheet_xls.cell_value(row, col)
    book_xlsx.save(dst_file_path)

# Работа со значениями и проверка на актуальность


def choose_article(n_file):
    wb = openpyxl.load_workbook(n_file)
    sheet = wb.active
    index = 16
    iB = 'B16'
    iD = 'D16'
    iI = 'I16'
    iA = 'A16'
    while sheet[iB].value:
        if sheet[iI].value == 'Создано' and sheet[iA].value:
            value = str(sheet[iB].value)
            article = value[::-1]
            article = article[:8]
            article = article[::-1]
            if article[1] == '\n':
                article = article[2:]
            article_list.append(article)
            amount.append(sheet[iD].value)
            index = index + 1
            iB = str('B'+str(index))
            iI = str('I' + str(index))
            iD = str('D' + str(index))
            iA = str('A' + str(index))
        else:
            index = index + 1
            iB = str('B' + str(index))
            iI = str('I' + str(index))
            iD = str('D' + str(index))
            iA = str('A' + str(index))

    print(article_list, amount)
    return article_list, amount

# Запись всех данных в конечный документ и зачистка от нежелательных данных


def write_in_file(n_file, article, amount):
    wb = openpyxl.load_workbook(n_file)
    sheet = wb.active
    # удаляем содержимое файла
    for row in range(0, 50):
        for col in range(0, 200):
            sheet.cell(row=row+1, column=col+1).value = ''
    # записываем нужные данные
    for a in range(len(article)):
        cell_a = sheet.cell(row=a+1, column=1)
        cell_a.value = int(article[a])
    for b in range(len(amount)):
        cell_b = sheet.cell(row=b+1, column=2)
        cell_b.value = amount[b]
    wb.save(n_file)


# создаем имя для нового файла
file_name = file_name('1.xls')
# создаем новый файл и копируем данные из загруженного файла формата xls в новый созданный файл xlxs
cvt_xls_to_xlsx('1.xls', file_name)
# выбираем необходимые значения со всем фильтрами и записываем их в два списка
article_list_data, amount_data = choose_article(file_name)
# записываем получившиеся списки в новый файл
write_in_file(file_name, article_list_data, amount_data)
